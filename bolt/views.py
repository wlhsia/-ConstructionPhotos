# 修改日期 2020/6/22
# 環境版本 
# CUDA:9.0
# cuDNN	v7.2.1 for CUDA 9.0
# tensorflow-gpu	1.12.0
# keras	2.0.6
# opencv-python	4.2.0
# pillow	5.2.0
# numpy	1.15.0
# numba	0.38.0
# scikit-image	0.14.2

from django.shortcuts import render
from .forms import FileFieldForm, DataForm
from web_project import settings
import os
import cv2
import numpy as np
import pandas as pd
from pdf2image import convert_from_path
import json
from sklearn.cluster import KMeans
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment,Border,Side
from openpyxl.drawing.image import Image as Image_xl
from PIL import Image
import itertools
from django.http import JsonResponse,HttpResponse
from functools import reduce
import shutil
from sklearn.externals import joblib
from bolt.models import feature, upload_record
from .unet import *
import keras
from numba import cuda
from datetime import datetime
import tempfile
from django.contrib.auth.decorators import login_required
# Create your views here.

# 設定路徑
pdfs_path = os.path.join(settings.BASE_DIR,'pdfs')
feature_path = os.path.join(settings.BASE_DIR,'feature')
model_path = os.path.join(settings.BASE_DIR,'model')
result_path = os.path.join(settings.BASE_DIR,'result')
imgs_path = os.path.join(settings.BASE_DIR,'imgs')

graph = tf.get_default_graph()
model = unet()
model.load_weights(os.path.join(model_path,'unet.hdf5'))

# 首頁
@login_required
def index(request):
    form = FileFieldForm()
    return render(request,'bolt/index.html',{'form':form })

# 上傳施工相片PDF
@login_required
def upload(request):
    if request.method == "POST":
        form = FileFieldForm(request.POST, request.FILES)
        files = request.FILES.getlist('file_field')
        if form.is_valid():
            # files = sorted(files)
            # pdf_path = os.path.join(settings.BASE_DIR,file_name[0])
            pdf_path = os.path.join(settings.BASE_DIR,files[0].name)
            if not os.path.isdir(pdf_path):
                os.mkdir(pdf_path)
            for file in files:
                upload_path = os.path.join(pdf_path,file.name)
                with open(upload_path,'wb+') as f:
                    for content in file.chunks():
                        f.write(content)
            pdf_list = os.listdir(pdf_path)
        else:
            pdf_list = []
    else:
        form = FileFieldForm()
        pdf_list = []
    form_data = DataForm()
    return render(request,'bolt/index.html', {'form':form, 'pdfs':pdf_list,'form_data':form_data})

# 刪除已上傳施工相片PDF
@login_required
def delete(request,dir,pdf):
    pdf_path = os.path.join(settings.BASE_DIR,dir)
    os.remove(os.path.join(pdf_path,pdf))
    pdf_list = os.listdir(pdf_path)
    if len(pdf_list)==0:
        shutil.rmtree(pdf_path)
    form = FileFieldForm()
    form_data = DataForm()
    return render(request, 'bolt/index.html', {'form':form, 'pdfs':pdf_list,'form_data':form_data})

# 下載比對結果
@login_required
def download(request,result):
    path = os.path.join(settings.BASE_DIR,'result')
    with open(os.path.join(path,result), 'rb') as f:
        response = HttpResponse(f)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename="'+result+'"'
    return response

# 比對
@login_required
def compare(request,dir):

    # 擷取上傳PDF
    pdf_path = os.path.join(settings.BASE_DIR,dir)
    pdfs=os.listdir(pdf_path)
    img_path = os.path.join(settings.BASE_DIR,dir+'img')
    if not os.path.isdir(img_path):
        os.mkdir(img_path)

    # 裁切施工相片PDF檔
    for pdf in pdfs:
        crop_img(pdf_path,pdf,img_path)
        shutil.move(os.path.join(pdf_path,pdf),os.path.join(pdfs_path,pdf))
    shutil.rmtree(pdf_path)
    # keras.backend.clear_session()
    # cuda.select_device(0)
    # cuda.close()

    # 將照片轉為特徵
    imgs_nm = os.listdir(img_path)
    dict_phash = {}
    dict_hist = {}
    for img_nm in imgs_nm:
        img=Image.open(os.path.join(img_path,img_nm))
        dict_phash[img_nm] = phash(img)
        dict_hist[img_nm] = hist(img)
        shutil.move(os.path.join(img_path,img_nm),os.path.join(imgs_path,img_nm))
    shutil.rmtree(img_path)

    df_result_1 = pd.DataFrame(columns = ['img1','img2','similary'])
    
    # 上傳相片倆倆比對
    for i,(img1_nm,img2_nm) in enumerate(itertools.combinations(imgs_nm, 2)):
        distance = bin(dict_phash[img1_nm] ^ dict_phash[img2_nm]).count('1')
        similary_phash = 1 - distance / 1024
        if similary_phash > 0.9:
            if len(df_result_1) == 0:
                df_result_1 = df_result_1.append({'img1':'上傳的相片','img2':'上傳的相片','similary':''},ignore_index=True)
            if len(df_result_1.query("img2 == '{}'".format(img1_nm))) == 0:
                df_result_1 = df_result_1.append({'img1':img1_nm,'img2':img2_nm,'similary':similary_phash},ignore_index=True)
           
    # 將上傳相片分群   
    dict_bin = {}
    for img_nm in imgs_nm:
        dict_bin[img_nm]=[bool(int(d)) for d in str(bin(dict_phash[img_nm]))[2:].zfill(1024)]
    kmeans = joblib.load(os.path.join(model_path,'KMeans_model.m'))
    pred = kmeans.predict(list(dict_bin.values()))
    dict_group = dict(zip(list(dict_bin.keys()), pred))
    df_clu = pd.DataFrame(data = list(dict_bin.keys()),columns = ['image'])
    df_clu['group'] = pred

    df_result_2 = pd.DataFrame(columns = ['img1','img2','similary'])

    # 上傳相片與相同群組的資料庫相片比對
    for group in set(df_clu.group):
        imgs_nm = df_clu.query("group == {}".format(group)).image
        imgs_data = feature.objects.filter(group=group)
        if imgs_data.exists():
            for img1_nm in imgs_nm:
                for img2_data in imgs_data:
                    img2_nm = img2_data.img_name
                    img2_phash = int(img2_data.phash)
                    distance = bin(dict_phash[img1_nm] ^ img2_phash).count('1')
                    similary_phash = 1 - distance / 1024
                    if similary_phash > 0.9:
                        if len(df_result_2) == 0:
                            df_result_1 = df_result_1.append({'img1':'資料庫的相片','img2':'上傳的相片','similary':''},ignore_index=True)
                        if len(df_result_2.query("img2 == '{}'".format(img1_nm))) == 0:
                            df_result_2 = df_result_2.append({'img1':img2_nm,'img2':img1_nm,'similary':similary_phash},ignore_index=True)

    df_result_2 = df_result_2.sort_values(by='img2')

    df_result = pd.concat([df_result_1,df_result_2], axis=0, ignore_index=True)

    # 製作顯示結果EXCEL表
    df_result.to_excel(os.path.join(result_path,'比對結果.xlsx'),engine = 'xlsxwriter',index = False)
    wb = load_workbook(filename= os.path.join(result_path,'比對結果.xlsx'))
    sht = wb['Sheet1']
    sht.insert_rows(1)
    sht.merge_cells('A1:E1')
    sht['A1'] = '工程案件相片重複性辨識'
    sht['A1'].font = Font(size=16, b=True, underline='single')
    sht['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sht.insert_rows(2)
    proj_num=dir.split('_')[0]
    sht['A2'] = '工程編號：'+ proj_num
    sht['E2'] = '日期：' + datetime.now().strftime("%Y/%m/%d")
    sht['A2'].font = Font(size=14, b=True)
    sht['E2'].font = Font(size=14, b=True)
    side = Side(border_style=None)
    sht['A3'].border = Border(left=side, right=side, top=side, bottom=side)
    sht['B3'].border = Border(left=side, right=side, top=side, bottom=side)
    sht['C3'].border = Border(left=side, right=side, top=side, bottom=side)
    sht.column_dimensions["A"].width = 35
    sht.column_dimensions["B"].width = 35
    sht.column_dimensions["D"].width = 25
    sht.column_dimensions["E"].width = 25

    # 顯示結果於網頁上
    if len(df_result) == 0:
        sht.insert_rows(2)
        sht['A5'] = '比對結果無重複相片'
        r = '比對結果無重複相片，是否將' + str(len(dict_phash)) + '張相片寫入資料庫?'
        form_data = DataForm(initial={"phash":str(dict_phash), "group": str(dict_group)})
    else:
        for img_nm in set(df_result.img2):
            if img_nm != '上傳的相片':
                del dict_phash[img_nm]
                del dict_group[img_nm]
        if len(dict_phash) !=0:
            r = '是否將無重複的'+ str(len(dict_phash)) +'張相片寫入資料庫?'
        else:
            r = '全部相片重複'
        form_data = DataForm(initial={"phash":str(dict_phash), "group": str(dict_group)})
        
        save_path = os.path.join(settings.BASE_DIR,'bolt/static/same_imgs')
        df_result['img1_path']=''
        df_result['img2_path']=''
        for idx,row in df_result.iterrows():
            if row.img2 != '上傳的相片':
                sht.row_dimensions[idx+4].height = 70
                img1_nm,img2_nm = row.img1,row.img2
                with Image.open(os.path.join(imgs_path,img1_nm)) as img:
                    img = img.resize((160,90),Image.ANTIALIAS)
                    img.save(os.path.join(save_path,img1_nm))
                with Image.open(os.path.join(imgs_path,img2_nm)) as img:
                    img = img.resize((160,90),Image.ANTIALIAS)
                    img.save(os.path.join(save_path,img2_nm))
                df_result.loc[idx,'img1_path']='/static/same_imgs/'+img1_nm
                df_result.loc[idx,'img2_path']='/static/same_imgs/'+img2_nm
                sht.add_image(Image_xl(os.path.join(save_path,img1_nm)),"D"+str(idx+4))
                sht.add_image(Image_xl(os.path.join(save_path,img2_nm)),"E"+str(idx+4))    

    wb.save(os.path.join(result_path,'比對結果.xlsx'))
    wb.close()
    result_list=os.listdir(result_path)
    form = FileFieldForm()
    pdf_list=[]

    upload_record.objects.create(user = request.user.username, pdf=dir, timestamp = datetime.now()) 

    return render(request, 'bolt/index.html', {'form':form, 'pdfs':pdf_list ,'results':result_list,'values':df_result.values.tolist(),'form_data': form_data,'r':r})

# 將相片寫入資料庫
@login_required
def save(request):
    form = DataForm()
    if request.method == 'POST':
        form = DataForm(request.POST, request.FILES)
        if form.is_valid():
            phash = request.POST['phash']
            group = request.POST['group']
            dict_phash = eval(phash)
            dict_group = eval(group)
            for img_nm in dict_phash.keys():
                data = feature.objects.filter(img_name = img_nm)
                if data.exists():
                    data.update(phash = dict_phash[img_nm])
                    data.update(group = int(dict_group[img_nm]))
                else:
                    feature.objects.create(img_name=img_nm,phash=dict_phash[img_nm],group = int(dict_group[img_nm]))
    form = FileFieldForm()
    pdf_list=[]
    form_data = DataForm()
    return render(request, 'bolt/index.html', {'form':form, 'pdfs':pdf_list,'form_data': form_data})


# 轉換pHash特徵
def phash(img):
    img = img.resize((32,32), Image.ANTIALIAS).convert('L')
    avg = reduce(lambda x, y: x + y, img.getdata()) / 1024.
    hash_value=reduce(lambda x, y: x | (y[1] << y[0]), enumerate(map(lambda i: 0 if i < avg else 1, img.getdata())), 0)
    return hash_value

# 轉換hist特徵
def hist(img):
    img = img.resize((256,256)).convert("L")
    hist_value = []
    for i in range(0,256,128):
        for j in range(0,256,128):
            sub_image = img.crop((i,j,i+128,j+128)).copy()
            hist_value.append((sub_image.histogram()))
    return hist_value


# 裁切PDF
def crop_img(pdf_path,pdf,img_path):
    # model = unet()
    # model.load_weights(os.path.join(model_path,'unet.hdf5'))
    with tempfile.TemporaryDirectory(dir= 'D:/temp') as path:
        page_imgs = convert_from_path(os.path.join(pdf_path,pdf), output_folder=path, dpi=600)
        for page_number,page_img in enumerate(page_imgs):
            if page_img.size[0] < page_img.size[1]:
                page_img = page_img.rotate(90,Image.NEAREST,expand =True)
            page_img = np.array(page_img)
            rgb = cv2.cvtColor(page_img,cv2.COLOR_BGR2RGB)
            hsv = cv2.cvtColor(page_img,cv2.COLOR_BGR2HSV)
            gray = cv2.cvtColor(page_img,cv2.COLOR_BGR2GRAY)
            #unet
            img = cv2.resize(gray, (256, 256))
            img = np.reshape(img,img.shape+(1,)) if (not False) else img
            img = np.reshape(img,(1,)+img.shape)
            with graph.as_default():
                result=model.predict(img)
            # result = model._make_predict_function(img)
            result=result[0]
            img = labelVisualize(2,COLOR_DICT,result) if False else result[:,:,0]
            img = cv2.resize(img, (page_img.shape[1],page_img.shape[0]))
            img = (img*255).astype(np.uint8)
            (thresh, im_bw) = cv2.threshold(img, 0.05, 255, 0)
            contours, hierarchy = cv2.findContours(im_bw, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            unet_conts = [contour for contour in contours if cv2.contourArea(contour)> 500000]
            #sat
            # 取出飽和度
            saturation = hsv[:,:,1]
            _, threshold = cv2.threshold(saturation, 1, 255.0, cv2.THRESH_BINARY)
            # 2值化圖去除雜訊
            kernel_radius = int(threshold.shape[1]/300)
            kernel = np.ones((kernel_radius, kernel_radius), np.uint8)
            threshold = cv2.morphologyEx(threshold,cv2.MORPH_OPEN,kernel)
            # 產生等高線
            contours, hierarchy = cv2.findContours(threshold, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            sat_conts = [contour for contour in contours if cv2.contourArea(contour)> 500000]

            if len(unet_conts) == 6:
                conts=unet_conts
            elif len(sat_conts) == 6:
                conts=sat_conts
            elif len(sat_conts) > len(unet_conts):
                conts=sat_conts
            else:
                conts=unet_conts

            sortY_conts = sorted([cont for cont in conts],key = lambda x:x[0][0][1],reverse=False)
            up_conts = sortY_conts[:3]
            up_conts = sorted([cont for cont in up_conts],key = lambda x:x[0][0][0],reverse=False)
            down_conts = sortY_conts[3:]
            down_conts = sorted([cont for cont in down_conts],key = lambda x:x[0][0][0],reverse=False)
            merge_conts = up_conts+down_conts

            for i,c in enumerate(merge_conts):
                # 嘗試在各種角度，以最小的方框包住面積最大的等高線區域，以紅色線條標示
                rect = cv2.minAreaRect(c)
                box = cv2.boxPoints(rect)
                box = np.int0(box) 
                angle = rect[2]
                if angle < -45:
                    angle = 90 + angle
                # 以影像中心為旋轉軸心
                (h, w) = page_img.shape[:2]
                center = (w // 2, h // 2)
                # 計算旋轉矩陣
                M = cv2.getRotationMatrix2D(center, angle, 1.0)
                # 旋轉圖片
                rotated = cv2.warpAffine(rgb, M, (w, h),flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_CONSTANT)
                # 旋轉紅色方框座標
                pts = np.int0(cv2.transform(np.array([box]), M))[0]
                #  計算旋轉後的紅色方框範圍
                y_min = min(pts[0][0], pts[1][0], pts[2][0], pts[3][0])
                y_max = max(pts[0][0], pts[1][0], pts[2][0], pts[3][0])
                x_min = min(pts[0][1], pts[1][1], pts[2][1], pts[3][1])
                x_max = max(pts[0][1], pts[1][1], pts[2][1], pts[3][1])
                # 裁切影像
                img_crop = rotated[x_min:x_max, y_min:y_max]
                page_num = (str(page_number+1)).zfill(3)
                dst_filenm = '_'.join([pdf,'Page'+page_num,str(i+1)])+'.jpg'
                cv2.imwrite(os.path.join(img_path,dst_filenm),img_crop)
